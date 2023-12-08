import openpyxl
import time , datetime
from config_strategys import all_vars

class Excels():
    def __init__(self,path,from_existing_excel:bool,contracts_dict_from_allVars={}) -> None:
        if from_existing_excel == True :
            self.wb=openpyxl.load_workbook(path)
            self.wb_path=path
        else : 
            #setting the new excel name to current datetime
            current_dateTime = datetime.datetime.now()
            self.wb_path=path+"\\"+str(current_dateTime.strftime('%Y.%m.%d_%H-%M-%S'))+".xlsx"
            self.wb =openpyxl.Workbook()
            self.create_strategy1_excel(contracts_dict_from_allVars)
        
        self.current_ws = self.wb.active

    def create_strategy1_excel(self,contracts_dict_from_allVars:dict):
        #creating a sheet for each contract
        for contract in contracts_dict_from_allVars :
            #creating new sheet for each contract 
            current_symbol=contract["symbol"]
            self.wb.create_sheet(current_symbol)
            #setting current wirking sheet
            self.current_ws = self.wb[current_symbol]
            #setting up initial data table cells
            self.current_ws["A1"] = "Date"
            self.current_ws["B1"] = "Price"
            self.current_ws["C1"] = "Action"
            self.current_ws["D1"] = "Signal"
            self.current_ws["E1"] = "Zscore"
            self.current_ws["F1"] = "emaZ"
            self.current_ws["G1"] = "real p&l"
            self.current_ws["H1"] = "compound p&l"
        
        #saving
        self.wb.save(self.wb_path)
    
    def write_line_strategy1(self,date,price,action,signal,zscore,emaz,realPnL,compoundPnL):
        current_row=self.current_ws.max_row+1

        self.current_ws[f"A{current_row}"] = date
        self.current_ws[f"B{current_row}"] = price
        self.current_ws[f"C{current_row}"] = action
        self.current_ws[f"D{current_row}"] = signal
        self.current_ws[f"E{current_row}"] = zscore
        self.current_ws[f"F{current_row}"] = emaz
        self.current_ws[f"G{current_row}"] = realPnL
        self.current_ws[f"H{current_row}"] = compoundPnL

        #saving
        self.wb.save(self.wb_path)

if __name__ == "__main__" :
    #RUN THIS TO CREATE NEW EXCEL FOR STRATEGY 
    path = all_vars["strategy 1"]["other_variables"]["live_excel_path"] # / "excle_path_live" / change path to what you want 
    new_excle = Excels(path,False)
    new_excle.create_strategy1_excel(all_vars["strategy 1"]["contracts"])